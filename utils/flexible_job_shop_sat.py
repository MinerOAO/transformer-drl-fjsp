#!/usr/bin/env python3
# Copyright 2010-2024 Google LLC
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# https://github.com/google/or-tools/blob/stable/examples/python/flexible_job_shop_sat.py

"""Solves a flexible jobshop problems with the CP-SAT solver.

A jobshop is a standard scheduling problem when you must sequence a
series of task_types on a set of machines. Each job contains one task_type per
machine. The order of execution and the length of each job on each
machine is task_type dependent.

The objective is to minimize the maximum completion time of all
jobs. This is called the makespan.
"""

# overloaded sum() clashes with pytype.

import collections
import time, os
from openpyxl import Workbook, load_workbook
from ortools.sat.python import cp_model


class SolutionPrinter(cp_model.CpSolverSolutionCallback):
    """Print intermediate solutions."""

    def __init__(self) -> None:
        cp_model.CpSolverSolutionCallback.__init__(self)
        self.__solution_count = 0

    def on_solution_callback(self) -> None:
        """Called at each new solution."""
        print(
            f"Solution {self.__solution_count}, time = {self.wall_time} s,"
            f" objective = {self.objective_value}"
        )
        self.__solution_count += 1


def flexible_jobshop(jobs, num_jobs, num_machines, time_limit=10) -> None:
    """solve a small flexible jobshop problem."""
    # Data part.
    # jobs = [  # task = (processing_time, machine_id)
    #     [  # Job 0
    #         [(3, 0), (1, 1), (5, 2)],  # task 0 with 3 alternatives
    #         [(2, 0), (4, 1), (6, 2)],  # task 1 with 3 alternatives
    #         [(2, 0), (3, 1), (1, 2)],  # task 2 with 3 alternatives
    #     ],
    #     [  # Job 1
    #         [(2, 0), (3, 1), (4, 2)],
    #         [(1, 0), (5, 1), (4, 2)],
    #         [(2, 0), (1, 1), (4, 2)],
    #     ],
    #     [  # Job 2
    #         [(2, 0), (1, 1), (4, 2)],
    #         [(2, 0), (3, 1), (4, 2)],
    #         [(3, 0), (1, 1), (5, 2)],
    #     ],
    # ]

    # num_jobs = len(jobs)
    all_jobs = range(num_jobs)

    # num_machines = 3
    all_machines = range(num_machines)

    # Model the flexible jobshop problem.
    model = cp_model.CpModel()

    horizon = 0
    for job in jobs:
        for task in job:
            max_task_duration = 0
            for alternative in task:
                max_task_duration = max(max_task_duration, alternative[0])
            horizon += max_task_duration

    print(f"Horizon = {horizon}")

    # Global storage of variables.
    intervals_per_resources = collections.defaultdict(list)
    starts = {}  # indexed by (job_id, task_id).
    presences = {}  # indexed by (job_id, task_id, alt_id).
    job_ends: list[cp_model.IntVar] = []

    # Scan the jobs and create the relevant variables and intervals.
    for job_id in all_jobs:
        job = jobs[job_id]
        num_tasks = len(job)
        previous_end = None
        for task_id in range(num_tasks):
            task = job[task_id]

            min_duration = task[0][0]
            max_duration = task[0][0]

            num_alternatives = len(task)
            all_alternatives = range(num_alternatives)

            for alt_id in range(1, num_alternatives):
                alt_duration = task[alt_id][0]
                min_duration = min(min_duration, alt_duration)
                max_duration = max(max_duration, alt_duration)

            # Create main interval for the task.
            suffix_name = f"_j{job_id}_t{task_id}"
            start = model.new_int_var(0, horizon, "start" + suffix_name)
            duration = model.new_int_var(
                min_duration, max_duration, "duration" + suffix_name
            )
            end = model.new_int_var(0, horizon, "end" + suffix_name)
            interval = model.new_interval_var(
                start, duration, end, "interval" + suffix_name
            )

            # Store the start for the solution.
            starts[(job_id, task_id)] = start

            # Add precedence with previous task in the same job.
            if previous_end is not None:
                model.add(start >= previous_end)
            previous_end = end

            # Create alternative intervals.
            if num_alternatives > 1:
                l_presences = []
                for alt_id in all_alternatives:
                    alt_suffix = f"_j{job_id}_t{task_id}_a{alt_id}"
                    l_presence = model.new_bool_var("presence" + alt_suffix)
                    l_start = model.new_int_var(0, horizon, "start" + alt_suffix)
                    l_duration = task[alt_id][0]
                    l_end = model.new_int_var(0, horizon, "end" + alt_suffix)
                    l_interval = model.new_optional_interval_var(
                        l_start, l_duration, l_end, l_presence, "interval" + alt_suffix
                    )
                    l_presences.append(l_presence)

                    # Link the primary/global variables with the local ones.
                    model.add(start == l_start).only_enforce_if(l_presence)
                    model.add(duration == l_duration).only_enforce_if(l_presence)
                    model.add(end == l_end).only_enforce_if(l_presence)

                    # Add the local interval to the right machine.
                    intervals_per_resources[task[alt_id][1]].append(l_interval)

                    # Store the presences for the solution.
                    presences[(job_id, task_id, alt_id)] = l_presence

                # Select exactly one presence variable.
                model.add_exactly_one(l_presences)
            else:
                intervals_per_resources[task[0][1]].append(interval)
                presences[(job_id, task_id, 0)] = model.new_constant(1)

        if previous_end is not None:
            job_ends.append(previous_end)

    # Create machines constraints.
    for machine_id in all_machines:
        intervals = intervals_per_resources[machine_id]
        if len(intervals) > 1:
            model.add_no_overlap(intervals)

    # Makespan objective
    makespan = model.new_int_var(0, horizon, "makespan")
    model.add_max_equality(makespan, job_ends)
    model.minimize(makespan)

    # Solve model.
    solver = cp_model.CpSolver()
    solution_printer = SolutionPrinter()
    solver.parameters.max_time_in_seconds = time_limit
    status = solver.solve(model, solution_printer)

    # Print final solution.
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        print(f"Optimal objective value: {solver.objective_value}")
        for job_id in all_jobs:
            print(f"Job {job_id}")
            for task_id, task in enumerate(jobs[job_id]):
                start_value = solver.value(starts[(job_id, task_id)])
                machine: int = -1
                task_duration: int = -1
                selected: int = -1
                for alt_id, alt in enumerate(task):
                    if solver.boolean_value(presences[(job_id, task_id, alt_id)]):
                        task_duration, machine = alt
                        selected = alt_id
                print(
                    f"  task_{job_id}_{task_id} starts at {start_value} (alt"
                    f" {selected}, machine {machine}, duration {task_duration})"
                )
    print(solver.response_stats())
    return solver.objective_value # Do not use best bound

def single_run(filename, time_limit=10):
    lines = []
    num_jobs = 0
    num_machines = 0
    jobs = []
    with open(filename) as file_object:
        first_line = file_object.readline().split()
        num_jobs = int(first_line[0])
        num_machines = int(first_line[1])
        lines = file_object.readlines()

    for line in lines:
        line = line.split()
        if len(line) <= 0:
            break
        num_opes = int(line[0])
        opes = []
        index = 1
        for i in range(num_opes):
            num_ava_mas = int(line[index])
            index += 1
            actions = []
            for j in range(num_ava_mas):
                mas_id = int(line[index]) - 1 # Dataset->[1, ...], Or-Tools[0, ...]
                proc_time = int(line[index + 1])
                actions.append((proc_time, mas_id))
                index += 2
            opes.append(actions)
        jobs.append(opes)
    # print(jobs)
    print(F"Num Jobs:{num_jobs}, Num Machines:{num_machines}")
    start_time = time.time()
    makespan = flexible_jobshop(jobs, num_jobs, num_machines, time_limit)
    last_time = time.time() - start_time
    return makespan, last_time

def main(data_path="./data_test/Public/"):
    test_files = os.listdir(data_path)
    test_files.sort(key=lambda x: x[:-4])

    str_time = time.strftime("%Y%m%d_%H%M%S", time.localtime(time.time()))
    result_path = f"./save/OR_Tools_{str_time}.xlsx"
    
    if os.path.exists(result_path):
        wb = load_workbook(result_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Makespan', 'Computation Time'])

    for filename in test_files:
        if filename.startswith("Behnke56"):
            makespan, computing_time = single_run(data_path + filename, 600)
            ws.append([filename, makespan, computing_time])
    wb.save(result_path)

    # writer = pd.ExcelWriter(f'{save_path}/makespan_{str_time}.xlsx') 
    # data = pd.DataFrame(torch.tensor(makespans).t().tolist(), columns=["file_name"])
    # data.to_excel(writer, sheet_name='Sheet1', index=False, startcol=i_rules + 1)
    # writer._save()
    # # writer.close()
    # data = pd.DataFrame(torch.tensor(times).t().tolist(), columns=[rule])
    # data.to_excel(writer_time, sheet_name='Sheet1', index=False, startcol=i_rules + 1)
    # writer_time._save()

# single_run("./data_test/Public/Mk05.fjs")
main()